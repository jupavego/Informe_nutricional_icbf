# -*- coding: utf-8 -*-
"""
Extrae las coordenadas de cada unidad de servicio y las cruza con el reporte
de seguimiento nutricional por Codigo UDS. Tambien cruza los CUPOS de cada
UDS para medir cobertura: beneficiarios unicos con toma nutricional frente
al cupo autorizado de la unidad.

Fuente coordenadas/zona: BD ANALISIS COBERTURA v21.xlsm, hoja 'COORDENADAS UDS'
        columna 19 LATITUD y columna 20 LONGITUD, ya en grados decimales.
Fuente cupos: ICBFCUEUnidadesServicio <fecha>.xlsx, columna 'Cupos UDS',
        solo unidades con Estado de la UDS = ACTIVA.

Se descarta todo lo que caiga fuera del recuadro de Antioquia: hay filas con
cero, con el signo perdido o con la coma mal interpretada.
"""
import openpyxl, csv, json, io, os
from collections import defaultdict

XLSM = (r"E:\REGIONAL ANTIOQUIA\GESTION SUPERVISION REGIONAL\VIDEOCONFERENCIAS"
        r"\7. 2026\8. AGOSTO\06-08-2026 ALERTA COBERTURA\BD ANALISIS COBERTURA v21.xlsm")
CUPOS_XLSX = (r"E:\REGIONAL ANTIOQUIA\GESTION SUPERVISION REGIONAL\VIDEOCONFERENCIAS"
              r"\7. 2026\8. AGOSTO\06-08-2026 ALERTA COBERTURA"
              r"\ICBFCUEUnidadesServicio 06082026.xlsx")
PROC = (r"E:\REGIONAL ANTIOQUIA\GESTION SUPERVISION REGIONAL\VIDEOCONFERENCIAS"
        r"\7. 2026\6. JUNIO\18-06-2026 SEGUIMIENTO NUTRICIONAL\REPS\_procesado")
OUT = os.path.dirname(os.path.abspath(__file__))
# la raiz del proyecto es la carpeta que contiene a codigo/; los recursos ya
# procesados viven aparte para no mezclarlos con el codigo
RAIZ = os.path.dirname(OUT)
REC = os.path.join(RAIZ, "recursos")
os.makedirs(REC, exist_ok=True)
OUT = REC

# recuadro de Antioquia con un margen prudente
LAT = (5.2, 9.1)
LON = (-77.3, -73.7)


def clave(c):
    """El codigo UDS aparece con y sin cero a la izquierda."""
    c = str(c or "").strip()
    return c.lstrip("0") or c


print("leyendo coordenadas ...", flush=True)
wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=True)
ws = wb["COORDENADAS UDS"]
it = ws.iter_rows(values_only=True)
next(it)
coord = {}
fuera = 0
for r in it:
    try:
        cod = r[0]
        la = float(r[19]); lo = float(r[20])
    except (TypeError, ValueError, IndexError):
        continue
    if not (LAT[0] <= la <= LAT[1] and LON[0] <= lo <= LON[1]):
        fuera += 1
        continue
    coord[clave(cod)] = (round(la, 5), round(lo, 5))
print(f"  coordenadas validas: {len(coord)}  ·  descartadas fuera del recuadro: {fuera}", flush=True)

# maestro de UDS: zona y comuna, util para el mapa
ws2 = wb["CUENTAME"]
it2 = ws2.iter_rows(values_only=True)
h2 = list(next(it2))
i_cod = h2.index("Código UDS")
i_zona = h2.index("Zona Ubicación UDS")
zona = {}
for r in it2:
    try:
        zona[clave(r[i_cod])] = (r[i_zona] or "").strip()
    except (TypeError, IndexError):
        pass
wb.close()
print(f"  maestro CUENTAME: {len(zona)} unidades", flush=True)

# --- cupos por UDS, para medir cobertura ---
print("leyendo cupos por UDS ...", flush=True)
wbc = openpyxl.load_workbook(CUPOS_XLSX, data_only=True, read_only=True)
wsc = wbc.active
itc = wsc.iter_rows(values_only=True)
hc = None
for r in itc:
    if r and r[0] == "Tipo de Unidad":
        hc = list(r)
        break
if hc is None:
    raise SystemExit("No se encontro la fila de encabezado ('Tipo de Unidad') en " + CUPOS_XLSX)
i_cod2 = hc.index("Código UDS")
i_cupos = hc.index("Cupos UDS")
i_estado = hc.index("Estado de la UDS")
cupos = {}
descartados = 0
for r in itc:
    if not r or r[i_cod2] is None:
        continue
    if (r[i_estado] or "").strip().upper() != "ACTIVA":
        descartados += 1
        continue
    try:
        c = int(r[i_cupos])
    except (TypeError, ValueError):
        continue
    if c > 0:
        cupos[clave(r[i_cod2])] = c
wbc.close()
print(f"  cupos validos: {len(cupos)} unidades activas  ·  descartadas (no activa): {descartados}", flush=True)

# --- cruce con el reporte nutricional ---
print("cruzando con el reporte ...", flush=True)
uds = {}
sin = defaultdict(int)
with io.open(os.path.join(PROC, "nn_ultima_toma.csv"), encoding="utf-8-sig") as fh:
    for r in csv.DictReader(fh):
        if r["estado"] != "VINCULADO":
            continue
        k = clave(r["cod_uds"])
        if k not in uds:
            cu = cupos.get(k)
            uds[k] = {"nom": r["uds"], "cz": r["cz"], "mun": r["municipio"],
                      "eas": r["eas"], "serv": r["servicio"], "n": 0,
                      "xy": coord.get(k), "z": zona.get(k, ""), "cupos": cu}
        uds[k]["n"] += 1
        if k not in coord:
            sin[r["cz"]] += 1

# --- cobertura POR TRIMESTRE: el lineamiento exige una toma por trimestre
# por beneficiario, no una ventana fija de meses sobre un corte que ademas
# varia segun que zonal vaya mas atrasado. Se toma el ultimo trimestre
# (ene-mar / abr-jun / jul-sep / oct-dic) que ya quedo COMPLETO en el
# corte actual, y se cuentan beneficiarios UNICOS con una toma ahi, por
# UDS, frente al cupo autorizado de esa UDS. ----------------------------
MESES_LARGO = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
               "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def trimestre_completo(mes_max):
    q = (mes_max - 1) // 3 + 1
    if mes_max % 3 != 0:
        q -= 1
    return None if q < 1 else [3 * q - 2, 3 * q - 1, 3 * q]


print("calculando cobertura del ultimo trimestre completo ...", flush=True)
NN_COMPLETO = os.path.join(PROC, "nn_completo.csv")
tmax_nn = 0
with io.open(NN_COMPLETO, encoding="utf-8-sig") as fh:
    for r in csv.DictReader(fh):
        try:
            tmax_nn = max(tmax_nn, int(r["toma"]))
        except (TypeError, ValueError):
            pass
TRIM = trimestre_completo(tmax_nn)
tri_docs = defaultdict(set)
if TRIM:
    TRIM_LBL = MESES_LARGO[TRIM[0]] + "-" + MESES_LARGO[TRIM[-1]]
    with io.open(NN_COMPLETO, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            try:
                t = int(r["toma"])
            except (TypeError, ValueError):
                continue
            if t in TRIM and r["estado"] == "VINCULADO":
                tri_docs[clave(r["cod_uds"])].add(r["doc"])
    print(f"  trimestre completo mas reciente: {TRIM_LBL} (meses {TRIM}, toma maxima {tmax_nn})", flush=True)
else:
    TRIM_LBL = None
    print(f"  aun no hay un trimestre completo (toma maxima {tmax_nn}); cobertura queda sin dato", flush=True)

for k, v in uds.items():
    v["n_trim"] = len(tri_docs.get(k, ()))
    v["cobertura"] = round(100.0 * v["n_trim"] / v["cupos"], 1) if v["cupos"] else None

con = sum(1 for v in uds.values() if v["xy"])
benef_con = sum(v["n"] for v in uds.values() if v["xy"])
benef_tot = sum(v["n"] for v in uds.values())
print(f"\nUDS en el reporte : {len(uds)}")
print(f"  con coordenada  : {con}  ({100.0*con/len(uds):.1f} %)")
print(f"  beneficiarios cubiertos: {benef_con:,} de {benef_tot:,}".replace(",", ".")
      + f"  ({100.0*benef_con/benef_tot:.1f} %)")
if sin:
    print("\nBeneficiarios en UDS sin coordenada, por centro zonal:")
    for k, v in sorted(sin.items(), key=lambda x: -x[1]):
        print(f"   {k:<28}{v:>7,}".replace(",", "."))

# --- cobertura por cupos: beneficiarios unicos del trimestre / cupo autorizado ---
con_cupo = [v for v in uds.values() if v["cupos"]]
sin_cupo = len(uds) - len(con_cupo)
if con_cupo and TRIM:
    prom = sum(v["cobertura"] for v in con_cupo) / len(con_cupo)
    sobre = sum(1 for v in con_cupo if v["cobertura"] > 100)
    print(f"\nCobertura por cupo UDS, trimestre {TRIM_LBL}  "
          f"({len(con_cupo)} unidades con cupo, {sin_cupo} sin cupo o sin cruce):")
    print(f"  cobertura promedio: {prom:.1f} %  ·  unidades por encima del cupo: {sobre}")

cupos_csv = os.path.join(PROC, "cobertura_uds.csv")
with io.open(cupos_csv, "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["cod_uds", "uds", "cz", "municipio", "eas", "servicio", "cupos",
                "beneficiarios_con_toma_historico", "beneficiarios_con_toma_trimestre",
                "cobertura_trimestre_pct"])
    filas = sorted(uds.items(), key=lambda kv: (kv[1]["cobertura"] is None, kv[1]["cobertura"] or 0))
    for k, v in filas:
        w.writerow([k, v["nom"], v["cz"], v["mun"], v["eas"], v["serv"],
                    v["cupos"] or "", v["n"], v["n_trim"],
                    v["cobertura"] if v["cobertura"] is not None else ""])
print(f"escrito {cupos_csv}  ·  {len(uds)} unidades")

salida = [{"c": k, "n": v["nom"][:44], "cz": v["cz"], "mu": v["mun"],
           "e": v["eas"][:46], "s": v["serv"][:44], "b": v["n"],
           "y": v["xy"][0], "x": v["xy"][1], "z": v["z"],
           "cu": v["cupos"], "bt": v["n_trim"], "cb": v["cobertura"]}
          for k, v in uds.items() if v["xy"]]
p = os.path.join(OUT, "uds.json")
json.dump(salida, io.open(p, "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
print(f"\nescrito {p}  {os.path.getsize(p)//1024} KB  ·  {len(salida)} unidades")

meta_p = os.path.join(OUT, "uds_trimestre.json")
json.dump({"tmax": tmax_nn, "meses": TRIM, "lbl": TRIM_LBL},
          io.open(meta_p, "w", encoding="utf-8"), ensure_ascii=False)
print(f"escrito {meta_p}")
